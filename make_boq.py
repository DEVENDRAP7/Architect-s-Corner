import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

random.seed(42)

# Realistic BOQ items grouped by trade. (description, unit, base_rate_range)
sections = {
    "EARTHWORK & SITE": [
        ("Site clearance and removal of vegetation", "Sqm", (15, 35)),
        ("Excavation in foundation in ordinary soil", "Cum", (180, 320)),
        ("Excavation in hard rock by chiselling", "Cum", (650, 950)),
        ("Earth filling in plinth with excavated soil", "Cum", (120, 200)),
        ("Sand filling in plinth, compacted", "Cum", (900, 1400)),
        ("Anti-termite treatment to foundation", "Sqm", (45, 90)),
        ("Disposal of surplus earth up to 1 km", "Cum", (90, 160)),
    ],
    "PCC & RCC CONCRETE": [
        ("PCC 1:4:8 in foundation", "Cum", (4200, 5200)),
        ("PCC 1:2:4 below flooring", "Cum", (4800, 5800)),
        ("RCC M20 in footings", "Cum", (6200, 7400)),
        ("RCC M25 in columns", "Cum", (6800, 8200)),
        ("RCC M25 in beams", "Cum", (6900, 8300)),
        ("RCC M25 in slab", "Cum", (6700, 8100)),
        ("RCC M30 in retaining wall", "Cum", (7400, 8900)),
        ("TMT steel reinforcement Fe500", "Kg", (62, 78)),
        ("Centering and shuttering for columns", "Sqm", (260, 380)),
        ("Centering and shuttering for slab", "Sqm", (240, 360)),
        ("Centering and shuttering for beams", "Sqm", (280, 400)),
    ],
    "MASONRY": [
        ("Brick masonry 230mm in CM 1:6", "Cum", (5200, 6400)),
        ("Brick masonry 115mm in CM 1:4", "Sqm", (520, 720)),
        ("AAC block masonry 200mm", "Cum", (4200, 5400)),
        ("AAC block masonry 100mm", "Sqm", (420, 600)),
        ("Stone masonry in foundation", "Cum", (3800, 4900)),
        ("RCC band/lintel over openings", "Rm", (320, 480)),
    ],
    "PLASTERING & FINISH": [
        ("Internal plaster 12mm CM 1:4", "Sqm", (180, 280)),
        ("External plaster 20mm double coat", "Sqm", (240, 360)),
        ("Ceiling plaster 6mm", "Sqm", (160, 240)),
        ("Waterproof plaster to sunk slabs", "Sqm", (260, 380)),
        ("Neeru/punning finish to walls", "Sqm", (60, 110)),
    ],
    "FLOORING & TILING": [
        ("Vitrified tile flooring 600x600", "Sqm", (650, 950)),
        ("Ceramic tile flooring 300x300", "Sqm", (420, 620)),
        ("Granite flooring 18mm", "Sqm", (1400, 2200)),
        ("Kota stone flooring", "Sqm", (520, 780)),
        ("Anti-skid tile in toilets", "Sqm", (480, 720)),
        ("Wall dado tiles up to 7ft", "Sqm", (520, 760)),
        ("Skirting 100mm height", "Rm", (90, 160)),
        ("Marble flooring polished", "Sqm", (1800, 2800)),
    ],
    "DOORS & WINDOWS": [
        ("Teak wood door frame 100x60mm", "Rm", (380, 560)),
        ("Flush door shutter 35mm both side laminate", "Sqm", (1600, 2400)),
        ("Panel door shutter teak", "Sqm", (3200, 4600)),
        ("UPVC sliding window 2 track", "Sqm", (1100, 1700)),
        ("Aluminium glazed window", "Sqm", (1300, 1900)),
        ("MS grill for windows", "Kg", (110, 160)),
        ("SS railing for staircase", "Rm", (1400, 2200)),
        ("Door hardware set (hinges, lock, handle)", "Set", (1200, 2400)),
    ],
    "PAINTING": [
        ("Cement primer to walls", "Sqm", (28, 48)),
        ("Two coats acrylic emulsion interior", "Sqm", (60, 100)),
        ("Two coats weatherproof exterior paint", "Sqm", (75, 120)),
        ("Enamel paint to MS/wood", "Sqm", (55, 95)),
        ("Texture paint to feature wall", "Sqm", (160, 280)),
        ("Putty 2 coats and sanding", "Sqm", (45, 80)),
    ],
    "PLUMBING & SANITARY": [
        ("CPVC pipe 15mm with fittings", "Rm", (140, 220)),
        ("CPVC pipe 25mm with fittings", "Rm", (190, 290)),
        ("PVC pipe 110mm soil line", "Rm", (260, 380)),
        ("PVC pipe 75mm waste line", "Rm", (180, 280)),
        ("EWC wall mounted with flush tank", "No", (6500, 11000)),
        ("Wash basin with pillar cock", "No", (3200, 5200)),
        ("Kitchen sink SS single bowl", "No", (3800, 6200)),
        ("CP shower with arm", "No", (1200, 2400)),
        ("Floor trap with grating", "No", (320, 560)),
        ("Overhead water tank 1000L", "No", (7500, 11000)),
        ("Gate valve 25mm brass", "No", (420, 720)),
    ],
    "ELECTRICAL": [
        ("Concealed PVC conduit wiring point", "Point", (650, 1050)),
        ("Power plug point 16A", "Point", (850, 1350)),
        ("AC point with dedicated circuit", "Point", (1400, 2200)),
        ("Modular switch 6A", "No", (180, 320)),
        ("MCB distribution board 8 way", "No", (2400, 3800)),
        ("Copper wire 2.5 sqmm per coil", "Coil", (2200, 3200)),
        ("LED panel light 18W", "No", (420, 720)),
        ("Ceiling fan point with regulator", "Point", (750, 1200)),
        ("Earthing pit with copper plate", "No", (3200, 5200)),
        ("Telephone/LAN point", "Point", (450, 750)),
    ],
    "WATERPROOFING": [
        ("Terrace waterproofing with membrane", "Sqm", (180, 320)),
        ("Toilet sunk waterproofing", "Sqm", (220, 360)),
        ("Brick bat coba on terrace", "Sqm", (260, 420)),
        ("Crystalline waterproofing to basement", "Sqm", (240, 400)),
    ],
    "MISCELLANEOUS": [
        ("Staircase MS handrail", "Rm", (650, 1050)),
        ("Pest control treatment", "Sqm", (18, 38)),
        ("Building cleaning on completion", "Sqm", (12, 28)),
        ("Scaffolding for external work", "Sqm", (45, 90)),
        ("Curing of concrete and masonry", "Sqm", (8, 20)),
    ],
}

# build master item list, expand with size/grade variants to reach 200+
variants = ["", " - Type A", " - Type B", " - Grade I", " - Grade II",
            " (Ground Floor)", " (First Floor)", " (Second Floor)"]

items = []  # (desc, unit, qty, base_rate)
sno = 0
for sec, lst in sections.items():
    for desc, unit, (lo, hi) in lst:
        # add base + a couple variants so list is long & realistic
        nvar = random.choice([2, 3, 3, 4])
        for v in random.sample(variants, nvar):
            base = random.randint(lo, hi)
            qty = random.choice([5,10,12,15,20,25,30,40,50,60,80,100,120,150,200,250,300,500])
            qty = round(qty * random.uniform(0.6, 1.8))
            items.append((sec, desc + v, unit, qty, base))

# trim/ensure >= 210
items = items[:230]
print("Total items:", len(items))

def write_book(path, contractor, rate_factor, drop_idx=None, lowball_idx=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"
    # title rows (real BOQ sheets have headers before the table)
    ws["A1"] = "BILL OF QUANTITIES - PROPOSED RESIDENTIAL BUILDING"
    ws["A2"] = "Contractor: " + contractor
    ws["A3"] = ""
    hdr = ["S.No", "Description of Item", "Unit", "Quantity", "Rate (INR)", "Amount (INR)"]
    hrow = 4
    ws.append([]) if False else None
    for c, h in enumerate(hdr, start=1):
        ws.cell(row=hrow, column=c, value=h)
    # style header
    fill = PatternFill("solid", fgColor="1F4E78")
    bold = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, 7):
        cell = ws.cell(row=hrow, column=c)
        cell.fill = fill; cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    r = hrow + 1
    for i, (sec, desc, unit, qty, base) in enumerate(items):
        if drop_idx and i in drop_idx:
            continue  # this contractor did not quote this item
        rate = round(base * rate_factor * random.uniform(0.95, 1.05))
        if lowball_idx and i in lowball_idx:
            rate = round(rate * 0.4)  # suspiciously low
        amount = qty * rate
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=desc)
        ws.cell(row=r, column=3, value=unit)
        ws.cell(row=r, column=4, value=qty)
        ws.cell(row=r, column=5, value=rate)
        ws.cell(row=r, column=6, value=amount)
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = border
        r += 1
    # widths
    widths = [8, 52, 10, 12, 14, 16]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + c)].width = w
    ws.cell(row=r + 1, column=2, value="TOTAL")
    ws.cell(row=r + 1, column=6, value="=SUM(F5:F%d)" % r)
    ws.cell(row=r + 1, column=2).font = Font(bold=True)
    ws.cell(row=r + 1, column=6).font = Font(bold=True)
    wb.save(path)
    print("Saved", path)

# Contractor A: baseline rates, quotes everything
write_book("D:/Architect's Corner/Contractor-A.xlsx", "ABC Constructions", 1.00)
# Contractor B: ~8% higher overall, MISSING 3 items, 2 suspiciously low items
write_book("D:/Architect's Corner/Contractor-B.xlsx", "XYZ Builders", 1.08,
           drop_idx={37, 88, 150}, lowball_idx={12, 99})
